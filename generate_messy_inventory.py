# Generates the deliberately "lived-in" corporate inventory workbook used as
# the Level 2 challenge: whoever (or whatever) attempts the fix must
# reverse-engineer the data model and business rules buried in the mess.
#
# Mess catalogue (each one is a real-world pattern):
#   - versioned filename with copy-paste suffix
#   - merged title block + warnings instead of metadata
#   - header row NOT in row 1, trailing spaces in headers
#   - mixed SKU formats, duplicate items with different spellings
#   - quantities as numbers AND as free text ("ca. 200", "12 Kartons")
#   - prices as numbers AND as text with German formatting
#   - formula column that breaks where text invaded numeric cells
#   - reorder flag: sometimes IF formula, sometimes hardcoded "JA!!"
#   - category separator rows + a subtotal row in the MIDDLE of the data
#   - color as semantics (red = discontinued, yellow = "klären") w/o legend
#   - hidden column with abandoned old prices
#   - notes column carrying undocumented business rules and a phone number
#   - second tab whose column layout silently shifts halfway down
#   - vendor tab with all contact info crammed into one cell
#   - stale "ALT (nicht benutzen!)" copy of the whole sheet, different columns
#   - hidden "NICHT ANFASSEN" price tab that the main sheet VLOOKUPs into
#   - empty "Tabelle3" nobody ever deleted

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="D9D9D9")
BLUE = PatternFill("solid", fgColor="DDEBF7")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006", bold=True)
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=16)
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))

wb = Workbook()

# ---------------------------------------------------------------- main sheet
ws = wb.active
ws.title = "Lager AKTUELL"

ws.merge_cells("A1:K1")
ws["A1"] = "NORDFIX Befestigungstechnik GmbH  –  LAGERBESTAND HALLE 2"
ws["A1"].font = TITLE
ws.merge_cells("A2:K2")
ws["A2"] = "Stand: 03.06.2026 (KW 23)   Verantwortlich: K. Brandt   !!! BITTE KEINE ZEILEN LÖSCHEN OHNE RÜCKSPRACHE !!!"
ws["A2"].font = RED_FONT
ws["A3"] = "(Mengen Stichprobe vom 28.05., Rest geschätzt -M.K.)"
ws["A3"].font = Font(italic=True, size=9)

HDR_ROW = 5
headers = ["Art.-Nr.", "Bezeichnung ", "Menge", "Einheit", "Lagerort",
           "Lieferant", "EK Preis ", "Wert", "Mind.Best.", "Nachbestellen?", "Bemerkung / Notizen"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(HDR_ROW, c, h)
    cell.font = BOLD
    cell.fill = GREY

# row tuples: (sku, name, qty, unit, loc, vendor, price, minstock, flag, note, fill)
# qty/price may be text. flag None => IF formula, else literal. fill: None/RED/YELLOW
SEP = "SEP"      # category separator row
SUBTOTAL = "SUB"  # the infamous mid-data subtotal

rows = [
    (SEP, "—  SCHRAUBEN  —"),
    ("SCH-0042", "Spanplattenschraube 4x40 verz.", 14500, "Stk", "R01-A", "Würth", 0.018, 5000, None, "", None),
    ("SCH-0043", "Spanplattenschraube 4x50 verz.", 9200, "Stk", "R01-A", "Würth", 0.021, 5000, None, "", None),
    ("sch-0044", "Spanplattenschr. 5x60", "ca. 3000", "Stk", "R01-B", "Würth", 0.034, 4000, "JA!!", "Karton aufgerissen, neu zählen!", YELLOW),
    ("SCH_0045 (neu)", "Terrassenschraube A2 5x50 TX25", 6800, "Stk", "R01-C", "Schrauben-Klaus", 0.11, 2000, None, "neuer Artikel seit 04/26", None),
    ("SCH-0046", "Sechskantschraube M8x60 8.8", 2340, "Stk", "R02-A", "Fabory", 0.09, 1000, None, "", None),
    ("SCH-0047", "Sechskantschraube M10x80 8.8", 980, "Stk", "R02-A", "Fabory", 0.16, 1000, None, "Lieferzeit aktuell 6 Wochen!!", None),
    ("SCH-0048", "Sechskant M12x100", 410, "Stk", "R02-B", "Fabory", "0,28 €", 500, "JA!!", "s. Mail v. Hr. Meier 12.05.", YELLOW),
    ("10455", "Holzbauschraube 8x200 Tellerkopf", 1250, "Stk", "R02-C", "Heco", 0.42, 800, None, "alte Artikelnr. aus AS400!", None),
    ("10456", "Holzbauschraube 8x240 Tellerkopf", 760, "Stk", "R02-C", "Heco", 0.51, 800, None, "", None),
    ("SCH-0051", "Trockenbauschraube 3,5x35 phosph.", 22000, "Stk", "R01-D", "Würth", 0.009, 10000, None, "", None),
    ("SCH-0052", "Trockenbauschraube 3.5x45", "18 Kartons", "Karton(1000)", "R01-D", "Würth", 0.011, 8000, None, "Einheit umgestellt 02/26 (K.B.)", None),
    (SUBTOTAL, "Zwischensumme Schrauben"),
    (SEP, "—  DÜBEL / ANKER  —"),
    ("DUE-0101", "Nylondübel 8x40", 8400, "Stk", "R03-A", "Fischer", 0.025, 3000, None, "", None),
    ("DUE-0102", "Nylondübel 10x50", 5100, "Stk", "R03-A", "Fischer", 0.04, 3000, None, "", None),
    ("DUE-0103", "Schwerlastanker M12", 340, "Stk", "R03-B", "Fischer", 1.85, 200, None, "", None),
    ("DUE-0104", "Schwerlastanker M16", 95, "Stk", "R03-B", "Fischer", 3.2, 150, None, "Hr. Wagner fragen vor Bestellung! 0171-2233448", None),
    ("DUE-0105", "Injektionsmörtel 300ml", 48, "Kartusche", "R03-C  (KÜHL!)", "Fischer", 7.9, 60, "JA!!", "MHD beachten!! 3 Stk abgelaufen 05/26", YELLOW),
    ("DUE-0106", "Gasbetondübel GB 10", 0, "Stk", "R03-A", "Tox", 0.31, 0, "nein", "AUSGELAUFEN - nicht nachbestellen (M.K. 09/24)", RED),
    (SEP, "—  MUTTERN / SCHEIBEN  —"),
    ("MUT-0201", "Sechskantmutter M8 verz.", 11000, "Stk", "R04-A", "Fabory", 0.012, 4000, None, "", None),
    ("MUT-0202", "Sechskantmutter M10 verz.", 7400, "Stk", "R04-A", "Fabory", 0.019, 4000, None, "", None),
    ("MUT-0203", "Sechskantmutter M12", 3100, "Stk", "R04-A", "Fabory", "0,031", 2000, None, "Preis prüfen - Liste v. 2024?", YELLOW),
    ("MUT-0204", "Hutmutter M8 A2", 850, "Stk", "R04-B", "Schrauben-Klaus", 0.14, 500, None, "", None),
    ("SHB-0301", "Unterlegscheibe M8 DIN125", 18500, "Stk", "R04-C", "Fabory", 0.004, 8000, None, "", None),
    ("SHB-0302", "Unterlegscheibe M10 DIN125", 12200, "Stk", "R04-C", "Fabory", 0.006, 8000, None, "", None),
    ("SHB-0302a", "Unterlegscheibe M10 DIN 125 (groß)", 2100, "Stk", "R04-C", "Fabory", 0.009, 1000, None, "Doppelt mit SHB-0302?? klären -J.", YELLOW),
    (SEP, "—  WERKZEUG / SONSTIGES  —"),
    ("WKZ-0401", "Bit-Set TX 32-tlg", 26, "Set", "R05-A", "Wera", 18.5, 10, None, "", None),
    ("WKZ-0402", "Bohrer-Kassette HSS 1-10mm", 14, "Set", "R05-A", "Bosch", 24.9, 10, None, "", None),
    ("WKZ-0403", "Akkuschrauber GSR 18V (Vorführgerät)", 3, "Stk", "Büro Hr. Brandt", "Bosch", 129.0, 0, "nein", "nur intern, NICHT verkaufen", RED),
    ("ALT-0001", "Kreuzschlitzschraube 4x40 (alt)", 6300, "Stk", "R09-?", "???", 0.015, 0, "nein", "Restbestand Übernahme Fa. Petersen 2019, abverkaufen", RED),
    ("DIV-0501", "Kabelbinder 200mm schwarz", "ca. 40 Beutel", "Beutel(100)", "R05-C", "Hellermann", 1.2, 30, None, "", None),
    ("DIV-0502", "Sprühöl 400ml", 72, "Dose", "R05-D GEFAHRGUT", "WD-40", 3.4, 50, None, "Lagerung: max 100 Dosen (Brandschutz!)", None),
]

r = HDR_ROW
for row in rows:
    r += 1
    if row[0] == SEP:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = ws.cell(r, 1, row[1])
        c.font = BOLD
        c.fill = BLUE
        continue
    if row[0] == SUBTOTAL:
        ws.cell(r, 2, row[1]).font = BOLD
        # hand-edited range, off by one at the top — classic
        ws.cell(r, 8, f"=SUM(H7:H{r-1})").font = BOLD
        ws.cell(r, 8).number_format = '#,##0.00 "€"'
        for col in range(1, 12):
            ws.cell(r, col).fill = GREY
        continue
    sku, name, qty, unit, loc, vendor, price, minstock, flag, note, fill = row
    ws.cell(r, 1, sku)
    ws.cell(r, 2, name)
    ws.cell(r, 3, qty)
    ws.cell(r, 4, unit)
    ws.cell(r, 5, loc)
    ws.cell(r, 6, vendor)
    pc = ws.cell(r, 7, price)
    if isinstance(price, float):
        pc.number_format = '#,##0.000 "€"'
    wc = ws.cell(r, 8, f"=C{r}*G{r}")  # breaks (#WERT!) where qty/price is text
    wc.number_format = '#,##0.00 "€"'
    ws.cell(r, 9, minstock)
    if flag is None:
        ws.cell(r, 10, f'=IF(C{r}<I{r},"JA","")')
    else:
        fc = ws.cell(r, 10, flag)
        if "JA" in str(flag):
            fc.font = RED_FONT
    ws.cell(r, 11, note)
    if fill:
        for col in range(1, 12):
            ws.cell(r, col).fill = fill
    for col in range(1, 12):
        ws.cell(r, col).border = THIN

# grand total pinned a few rows below, with a label in the wrong column
r += 3
ws.cell(r, 6, "GESAMTWERT LAGER:").font = TITLE
gt = ws.cell(r, 8, f"=SUM(H6:H{r-3})")
gt.font = TITLE
gt.number_format = '#,##0.00 "€"'
ws.cell(r + 1, 6, "(ohne Halle 1! siehe alte Datei auf Q:\\Lager\\2023\\)").font = Font(italic=True, size=9)

# hidden column L: abandoned 2023 prices someone "kept just in case"
ws.cell(HDR_ROW, 12, "EK alt (2023) NICHT MEHR PFLEGEN")
ws.column_dimensions["L"].hidden = True

widths = [14, 36, 11, 13, 16, 15, 11, 12, 10, 14, 45]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------------------------- order log tab
ol = wb.create_sheet("Bestellungen 2026")
ol["A1"] = "Bestellungen 2026 (nur Halle 2!)"
ol["A1"].font = TITLE
hdr1 = ["Datum", "Art.-Nr.", "Menge", "Lieferant", "bestellt von", "geliefert?"]
for c, h in enumerate(hdr1, 1):
    ol.cell(3, c, h).font = BOLD
    ol.cell(3, c).fill = GREY
orders_old = [
    ("06.01.2026", "SCH-0042", 20000, "Würth", "KB", "ja"),
    ("19.01.2026", "DUE-0103", 500, "Fischer", "KB", "ja"),
    ("02.02.2026", "MUT-0201", 10000, "Fabory", "MK", "ja"),
    ("17.2.", "SHB-0301", 20000, "Fabory", "MK", "ja"),
    ("Anfang März", "WKZ-0401", 20, "Wera", "JS", "ja (1 defekt)"),
]
r = 3
for o in orders_old:
    r += 1
    for c, v in enumerate(o, 1):
        ol.cell(r, c, v)
# someone inserted a "Bestell-Nr." column in mid-March and re-headered mid-table
r += 2
hdr2 = ["Datum", "Bestell-Nr.", "Art.-Nr.", "Menge", "Lieferant", "bestellt von", "geliefert?"]
for c, h in enumerate(hdr2, 1):
    ol.cell(r, c, h).font = BOLD
    ol.cell(r, c).fill = YELLOW
orders_new = [
    ("23.03.2026", "B-2026-031", "SCH-0047", 2000, "Fabory", "KB", "ja"),
    ("08.04.2026", "B-2026-038", "DUE-0105", 100, "Fischer", "KB", "TEILLIEFERUNG 60 Stk!"),
    ("2026-04-29", "B-2026-044", "SCH_0045 (neu)", 10000, "Schrauben-Klaus", "JS", "ja"),
    ("12.05.2026", "B-2026-051", "SCH-0048", 1000, "Fabory", "KB", "offen"),
]
for o in orders_new:
    r += 1
    for c, v in enumerate(o, 1):
        ol.cell(r, c, v)
for i, w in enumerate([13, 13, 16, 9, 16, 12, 20], 1):
    ol.column_dimensions[get_column_letter(i)].width = w

# --------------------------------------------------------------- vendors tab
vd = wb.create_sheet("Lieferanten")
vd["A1"] = "Lieferanten & Konditionen"
vd["A1"].font = TITLE
for c, h in enumerate(["Firma", "Kontakt (alles)", "Rabatt", "Bemerkung"], 1):
    vd.cell(3, c, h).font = BOLD
    vd.cell(3, c).fill = GREY
vendors = [
    ("Würth", "Hr. Petersen 0171-4456120 lars.petersen@wuerth-nl-hh.de / Vertretung: Fr. Aydin", "Staffel s. Preisliste", "Rahmenvertrag bis 12/2026"),
    ("Fabory", "service@fabory.com Hotline 0800-3322874, Kundennr. 449-2231-K", "8%", "Lieferzeiten seit 03/26 schlecht!"),
    ("Fischer", "Hr. Wagner (Außendienst) 0171-2233448", "12% ab 500€", "IMMER Hr. Wagner direkt anrufen, Webshop-Preise stimmen nicht"),
    ("Schrauben-Klaus ", "Klaus 04101-55872 (vormittags)", "-", "Barzahlung Skonto 2%, Kleinmengen ok"),
    ("Heco", "über Großhandel Petersen & Co", "?", "Konditionen klären -KB"),
    ("Wera / Bosch", "über Werkzeug-Zentrale Nord, Fr. Schulz", "Liste 2024", "Preisliste 2026 anfordern!!"),
]
r = 3
for v in vendors:
    r += 1
    for c, val in enumerate(v, 1):
        vd.cell(r, c, val)
for i, w in enumerate([18, 60, 16, 38], 1):
    vd.column_dimensions[get_column_letter(i)].width = w

# ----------------------------------------------- stale copy nobody dares delete
old = wb.create_sheet("Lager ALT nicht benutzen!")
old["A1"] = "LAGERLISTE  (Übernahme aus AS400, Stand 11/2023)"
old["A1"].font = TITLE
old["A2"] = "ACHTUNG: veraltet!! aktuelle Liste siehe Blatt 'Lager AKTUELL'"
old["A2"].font = RED_FONT
# different column order and old numeric SKUs — the migration nobody finished
for c, h in enumerate(["Nr", "Menge", "Artikel", "Platz", "Preis DM/EUR umgerechnet"], 1):
    old.cell(4, c, h).font = BOLD
legacy = [
    (10452, 9000, "Spanplattenschraube 4x40", "A-01", 0.016),
    (10453, 4000, "Spanplattenschraube 4x50", "A-01", 0.019),
    (10455, 2200, "Holzbauschraube 8x200", "B-04", 0.39),
    (10460, 15000, "Kreuzschlitz 4x40", "A-02", 0.015),
    (10471, 7000, "Dübel 8mm", "C-01", 0.02),
]
r = 4
for row in legacy:
    r += 1
    for c, v in enumerate(row, 1):
        old.cell(r, c, v)
old.sheet_properties.tabColor = "FF0000"

# ------------------------------------------------ hidden price tab + empty tab
pz = wb.create_sheet("Preise NICHT ANFASSEN")
pz["A1"] = "VK-Kalkulation (Basis für Angebote!) - nur K. Brandt"
pz["A1"].font = RED_FONT
for c, h in enumerate(["Art.-Nr.", "EK", "Faktor", "VK netto", "VK brutto"], 1):
    pz.cell(3, c, h).font = BOLD
calc = [("SCH-0042", 0.018, 2.8), ("SCH-0046", 0.09, 2.4), ("DUE-0103", 1.85, 2.1),
        ("WKZ-0401", 18.5, 1.6), ("DIV-0502", 3.4, 2.2)]
r = 3
for sku, ek, f in calc:
    r += 1
    pz.cell(r, 1, sku)
    pz.cell(r, 2, ek)
    pz.cell(r, 3, f)
    pz.cell(r, 4, f"=B{r}*C{r}")
    pz.cell(r, 5, f"=D{r}*1.19")
pz.sheet_state = "hidden"

wb.create_sheet("Tabelle3")  # the eternal empty sheet

out = "Lagerbestand_Halle2_2026_FINAL_v3 (2).xlsx"
wb.save(out)
print("written:", out)
